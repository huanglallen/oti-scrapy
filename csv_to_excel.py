import pandas as pd
from openpyxl import load_workbook, Workbook

# File paths
csv_filename = 'genscript.csv'
sheet_name = csv_filename.replace('.csv', '')
excel_path = 'scrape-data-protein.xlsx'

# Load CSV data
df = pd.read_csv(csv_filename)

try:
    # Try to load the existing workbook
    wb = load_workbook(excel_path)
except FileNotFoundError:
    # Create new workbook if file doesn't exist
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

# Remove sheet if it already exists
if sheet_name in wb.sheetnames:
    std = wb[sheet_name]
    wb.remove(std)

# Create a new sheet with the desired name
ws = wb.create_sheet(title=sheet_name)

# Write headers
for col_num, column_title in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_num, value=column_title)

# Write data rows
for row_num, row_data in enumerate(df.values, start=2):
    for col_num, cell_value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max_length + 2

# Save workbook
wb.save(excel_path)
print("Output file created/updated in sheet:", sheet_name)
